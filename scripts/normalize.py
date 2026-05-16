"""
讀取出隊 Excel（all in one 直企格式），寫入 SQLite，並生成 P1/P2 截圖。
資料來源：直企P1（出隊資訊）、直企P2（隊員名單、留守資料）。
用法：
  python3 scripts/normalize.py data/raw/xxx.xlsx
  python3 scripts/normalize.py data/raw/          # 處理目錄下所有 xlsx
"""
import re
import sys
import sqlite3
import subprocess
import tempfile
from pathlib import Path

import fitz
import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image, ImageOps

DB_PATH         = Path(__file__).parent.parent / "db" / "sttmount.db"
STATIC_MAPS     = Path(__file__).parent.parent / "app" / "static" / "maps"
STATIC_PREVIEWS = Path(__file__).parent.parent / "app" / "static" / "previews"
RAW_DIR         = Path(__file__).parent.parent / "data" / "raw"
GPX_DIR         = Path(__file__).parent.parent / "app" / "static" / "gpx"

GPX_EXTS    = {".gpx", ".kml"}
MAP_EXTS    = {".pdf"}
RECORD_EXTS = {".txt", ".md"}

COUNTY_NORMALIZE = {
    "臺北市": "台北", "台北市": "台北",
    "新北市": "新北",
    "基隆市": "基隆",
    "宜蘭縣": "宜蘭",
    "桃園市": "桃園",
    "新竹市": "新竹", "新竹縣": "新竹",
    "苗栗縣": "苗栗",
    "臺中市": "台中", "台中市": "台中",
    "花蓮縣": "花蓮",
    "彰化縣": "彰化",
    "南投縣": "南投",
    "雲林縣": "雲林",
    "嘉義市": "嘉義", "嘉義縣": "嘉義",
    "臺南市": "台南", "台南市": "台南",
    "高雄市": "高雄",
    "屏東縣": "屏東",
    "臺東縣": "台東", "台東縣": "台東",
}

# P2 欄 A 的角色縮寫對應
ROLE_MAP = {"領": "領隊", "嚮": "嚮導", "隊": "隊員", "新": "新生"}


def roc_to_iso(text: str) -> str | None:
    """'民國 115 年 4 月 30 日 ...' → '2026-04-30'"""
    m = re.search(r"(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", str(text))
    if not m:
        return None
    y = int(m.group(1)) + 1911
    mo, d = int(m.group(2)), int(m.group(3))
    return f"{y:04d}-{mo:02d}-{d:02d}"


def extract_county_region(location: str):
    """'入山：臺東縣達仁鄉' → ('台東', '達仁')"""
    county = None
    for official, display in COUNTY_NORMALIZE.items():
        if official in location:
            county = display
            break
    region = None
    m2 = re.search(r"[縣市](.{1,6}?)[鄉鎮市區]", location)
    if m2:
        region = m2.group(1)
    return county, region


def capture_sheet_range(xlsx_path: Path, sheet_name: str, cell_range: str, output_path: Path):
    """將指定 sheet 的 cell range 截圖存為 PNG。"""
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[sheet_name]
        ws.print_area = cell_range
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        for name in list(wb.sheetnames):
            if name != sheet_name:
                del wb[name]
        tmp_xlsx = tmp / "preview.xlsx"
        wb.save(tmp_xlsx)
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             str(tmp_xlsx), "--outdir", str(tmp)],
            capture_output=True, timeout=60
        )
        if result.returncode != 0:
            print(f"    ⚠ LibreOffice 失敗：{result.stderr.decode()[:200]}")
            return
        pdf_path = tmp / "preview.pdf"
        if not pdf_path.exists():
            print(f"    ⚠ PDF 未生成")
            return
        doc = fitz.open(pdf_path)
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        pix.save(str(output_path))


def trim_whitespace(img: Image.Image, padding: int = 15) -> Image.Image:
    gray = img.convert("L")
    bbox = ImageOps.invert(gray).getbbox()
    if not bbox:
        return img
    l, t, r, b = bbox
    return img.crop((max(0, l - padding), max(0, t - padding),
                     min(img.width, r + padding), min(img.height, b + padding)))


def build_a4_preview(paths: list[Path], output_path: Path):
    A4_W, A4_H, GAP = 1240, 1754, 16
    imgs = [trim_whitespace(Image.open(p)) for p in paths if p.exists()]
    if not imgs:
        return
    target_w = A4_W
    resized = [img.resize((target_w, round(img.height * target_w / img.width)), Image.LANCZOS)
               for img in imgs]
    total_h = sum(img.height for img in resized) + GAP * (len(resized) - 1)
    if total_h > A4_H:
        scale = A4_H / total_h
        target_w = round(A4_W * scale)
        resized = [img.resize((target_w, round(img.height * target_w / img.width)), Image.LANCZOS)
                   for img in imgs]
        total_h = sum(img.height for img in resized) + GAP * (len(resized) - 1)
    canvas = Image.new("RGB", (max(img.width for img in resized), total_h), "white")
    y = 0
    for img in resized:
        canvas.paste(img, (0, y))
        y += img.height + GAP
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(str(output_path))


def scan_static_files(xlsx_path: Path, exp_id: int, conn: sqlite3.Connection):
    exp_folder = xlsx_path.parent
    if exp_folder == RAW_DIR:
        return

    folder_name = exp_folder.name

    gpx_file = GPX_DIR / f"{folder_name}.gpx"
    if gpx_file.exists():
        conn.execute(
            "INSERT OR IGNORE INTO gpx_files(expedition_id, filename, file_path) VALUES (?, ?, ?)",
            (exp_id, gpx_file.name, gpx_file.name),
        )

    map_pdf = STATIC_MAPS / f"{folder_name}.pdf"
    if map_pdf.exists():
        conn.execute(
            "INSERT OR IGNORE INTO map_files(expedition_id, file_path) VALUES (?, ?)",
            (exp_id, map_pdf.name),
        )

    rec_dir = exp_folder / "records"
    if rec_dir.is_dir():
        for f in sorted(rec_dir.iterdir()):
            if f.suffix.lower() in RECORD_EXTS:
                exists = conn.execute(
                    "SELECT 1 FROM records WHERE expedition_id=? AND filename=?",
                    (exp_id, f.name),
                ).fetchone()
                if not exists:
                    content = f.read_text(encoding="utf-8", errors="replace")
                    conn.execute(
                        "INSERT INTO records(expedition_id, filename, content) VALUES (?, ?, ?)",
                        (exp_id, f.name, content),
                    )

    conn.commit()
    print(f"    靜態檔案已掃描：{folder_name}/")


def parse_p1(ws):
    """從 直企P1(列印) 萃取出隊資訊。"""
    name = str(ws["D2"].value or "").strip() or None
    date_start = roc_to_iso(ws["C3"].value or "")
    date_end = roc_to_iso(ws["C4"].value or "")
    entry_loc = str(ws["F3"].value or "")
    exit_loc  = str(ws["F4"].value or "")
    county, region = extract_county_region(entry_loc)
    county_exit, region_exit = extract_county_region(exit_loc)
    return name, date_start, date_end, county, region, county_exit, region_exit


def parse_p2(ws):
    """從 直企P2(列印) 萃取留守資料（→ description）和隊員名單。"""
    # 留守資料：M/N 欄 rows 3–11
    desc_parts = []
    for r in range(3, 12):
        label = str(ws.cell(r, 13).value or "").strip()   # col M
        value = str(ws.cell(r, 14).value or "").strip()   # col N
        if label and value:
            desc_parts.append(f"{label}：{value}")
    garmin = str(ws["D10"].value or "").strip()
    if garmin.startswith("http"):
        desc_parts.append(f"Garmin 追蹤：{garmin}")
    notes = str(ws["D11"].value or "").strip()
    if notes:
        desc_parts.append(f"注意事項：{notes}")
    description = "\n".join(desc_parts) or None

    # 人員名單：rows 16+
    # col A = 角色縮寫（carry-forward）
    # col B(2) = 系級（第一行，格式「土木114\nE64102038」）
    # col D(4) = 姓名
    # col F(6) = 資歷（格式「A/奇萊東稜下嵐山」）
    members: list[tuple[str, str | None, str | None, str | None]] = []
    current_role: str | None = None
    for r in range(16, ws.max_row + 1):
        role_abbr = str(ws.cell(r, 1).value or "").strip()
        dept_raw  = str(ws.cell(r, 2).value or "").strip()
        name_raw  = str(ws.cell(r, 4).value or "").strip()
        exp_raw   = str(ws.cell(r, 6).value or "").strip()

        if role_abbr in ROLE_MAP:
            current_role = ROLE_MAP[role_abbr]

        name = name_raw.split("\n")[0].strip()
        department = dept_raw.split("\n")[0].strip() or None
        experience = exp_raw.split("\n")[0].strip() or None

        if name and current_role is not None:
            members.append((name, current_role, department, experience))

    return description, members


def normalize(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "直企P1(列印)" not in wb.sheetnames:
        print(f"  ⚠ 找不到「直企P1(列印)」sheet，跳過 {xlsx_path.name}")
        return

    ws_p1 = wb["直企P1(列印)"]
    name, date_start, date_end, county, region, county_exit, region_exit = parse_p1(ws_p1)

    if not name:
        print(f"  ⚠ 無法取得出隊名稱，跳過 {xlsx_path.name}")
        return
    if not date_start:
        print(f"  ⚠ 無法解析 date_start，跳過 {xlsx_path.name}")
        return

    description, members = None, []
    if "直企P2(列印)" in wb.sheetnames:
        description, members = parse_p2(wb["直企P2(列印)"])
    else:
        print(f"  ⚠ 找不到「直企P2(列印)」sheet，隊員與留守資料略過")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    existing = conn.execute(
        "SELECT id FROM expeditions WHERE name=? AND date_start=?",
        (name, date_start)
    ).fetchone()

    if existing:
        print(f"  → 已存在（id={existing[0]}）：{name}，跳過")
        conn.close()
        return

    cur = conn.execute(
        "INSERT INTO expeditions(name, date_start, date_end, county, region, region_exit, description) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (name, date_start, date_end, county, region, region_exit, description),
    )
    conn.commit()
    exp_id = cur.lastrowid

    for mname, mrole, mdept, mexp in members:
        conn.execute(
            "INSERT INTO members(expedition_id, name, role, department, experience) "
            "VALUES (?, ?, ?, ?, ?)",
            (exp_id, mname, mrole, mdept, mexp),
        )

    for c in {county, county_exit} - {None}:
        conn.execute(
            "INSERT OR IGNORE INTO expedition_counties(expedition_id, county) VALUES (?,?)",
            (exp_id, c),
        )
    conn.commit()

    scan_static_files(xlsx_path, exp_id, conn)
    conn.close()

    print(f"  ✓ 已插入：{name}（id={exp_id}）")
    print(f"    日期：{date_start} ～ {date_end or '—'}")
    print(f"    地點：{county or '—'} · {region or '—'}")
    print(f"    隊員：{len(members)} 人")

    # 生成截圖並合併
    STATIC_PREVIEWS.mkdir(parents=True, exist_ok=True)
    preview_path = STATIC_PREVIEWS / f"{exp_id}.png"
    with tempfile.TemporaryDirectory() as _tmp:
        _tmp = Path(_tmp)
        p1_path = _tmp / "p1.png"
        p2_path = _tmp / "p2.png"

        print(f"    截圖 P1...", end=" ", flush=True)
        if "直企P1(列印)" in wb.sheetnames:
            capture_sheet_range(xlsx_path, "直企P1(列印)", "A2:G27", p1_path)
            print("完成" if p1_path.exists() else "失敗")
        else:
            print("跳過")

        print(f"    截圖 P2...", end=" ", flush=True)
        if "直企P2(列印)" in wb.sheetnames:
            capture_sheet_range(xlsx_path, "直企P2(列印)", "B2:O11", p2_path)
            print("完成" if p2_path.exists() else "失敗")
        else:
            print("跳過")

        build_a4_preview([p1_path, p2_path], preview_path)

    if preview_path.exists():
        rel = f"previews/{exp_id}.png"
        conn = sqlite3.connect(DB_PATH)
        conn.execute("UPDATE expeditions SET preview_image=? WHERE id=?", (rel, exp_id))
        conn.commit()
        conn.close()
        print(f"    預覽圖：{rel}")


def main():
    if len(sys.argv) < 2:
        print("用法：python3 scripts/normalize.py <檔案.xlsx 或目錄>")
        sys.exit(1)

    target = Path(sys.argv[1])
    files = sorted(target.glob("**/*.xlsx")) if target.is_dir() else [target]

    for f in files:
        print(f"\n處理：{f.name}")
        try:
            normalize(f)
        except Exception as e:
            print(f"  ✗ 錯誤：{e}")


if __name__ == "__main__":
    main()
